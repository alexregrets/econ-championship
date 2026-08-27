"""Выгрузка датасета раунда: CSV, XLSX и словарь данных.

Главный тест здесь — не про формат, а про утечку. Если в выгрузку попадёт
истина (`CompanyGroundTruth`, настоящие `a` и `b`, пофирменные издержки),
турнир кончится: команде незачем оценивать то, что ей выдали числом. Такая
ошибка обнаруживается последней, потому что снаружи выгрузка выглядит
нормально, — поэтому проверяется первой.
"""

from __future__ import annotations

import csv
import io
from collections.abc import AsyncIterator

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import create_async_engine
from sqlmodel import SQLModel
from sqlmodel.ext.asyncio.session import AsyncSession

from core.cases import REGIME_COLUMN, supported_methods
from db import models  # noqa: F401  (регистрирует таблицы)
from db import repositories as repo
from db.enums import Method
from devshell.role_seed import seed_oil_2013
from devshell.seed import seed
from services.dataset_export import (
    DISCLAIMER,
    FORBIDDEN_COLUMNS,
    build_round_dataset,
    data_dictionary_markdown,
    to_csv,
    to_xlsx,
)


@pytest_asyncio.fixture
async def session() -> AsyncIterator[AsyncSession]:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.create_all)
    async with AsyncSession(engine, expire_on_commit=False) as s:
        yield s
    await engine.dispose()


# --------------------------------------------------------------------------- #
# Истина наружу не уходит
# --------------------------------------------------------------------------- #


async def test_export_never_contains_ground_truth_values(
    session: AsyncSession,
) -> None:
    """Ни истинных a и b, ни себестоимости — ни в одной ячейке.

    Сравниваются **значения ячеек**, а не текст файла. Поиск подстрокой даёт
    ложные срабатывания: «364» находится внутри цены 305.766364783, и такой
    тест краснел бы на здоровой выгрузке, а от этого его быстро ослабили бы
    до бесполезного.
    """
    summary = await seed_oil_2013(session)
    round_ = await repo.get_round(session, summary.round_id)
    assert round_ is not None
    secrets = (round_.market_a, round_.market_b, round_.market_mc)

    text = to_csv(await build_round_dataset(session, summary.round_id))
    rows = list(csv.DictReader(io.StringIO(text)))

    for row in rows:
        for value in row.values():
            for secret in secrets:
                assert abs(float(value) - secret) > 1e-6, (
                    f"в выгрузку утекло {secret}"
                )


async def test_export_has_no_forbidden_columns(session: AsyncSession) -> None:
    summary = await seed(session)
    dataset = await build_round_dataset(session, summary.round_id)

    names = {c.name.lower() for c in dataset.columns}
    assert not (names & FORBIDDEN_COLUMNS), f"запрещённые столбцы: {names}"


async def test_export_columns_and_rows_agree(session: AsyncSession) -> None:
    """Строка не может нести поле, которого нет в словаре: иначе студент
    получает столбец без объяснения, что это такое."""
    summary = await seed(session)
    dataset = await build_round_dataset(session, summary.round_id)

    declared = {c.name for c in dataset.columns}
    for row in dataset.rows:
        assert set(row) == declared


# --------------------------------------------------------------------------- #
# Подпись про учебные данные
# --------------------------------------------------------------------------- #


async def test_disclaimer_is_in_dictionary_and_workbook(
    session: AsyncSession,
) -> None:
    """Названия компаний настоящие, цифры выдуманные — это обязано быть сказано."""
    summary = await seed(session)
    dataset = await build_round_dataset(session, summary.round_id)

    assert DISCLAIMER in data_dictionary_markdown(dataset)

    book = load_workbook(io.BytesIO(to_xlsx(dataset)))
    dictionary_sheet = book["Словарь"]
    text = "\n".join(
        str(cell.value)
        for row in dictionary_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert DISCLAIMER in text


async def test_dictionary_describes_every_column(session: AsyncSession) -> None:
    summary = await seed(session)
    dataset = await build_round_dataset(session, summary.round_id)
    markdown = data_dictionary_markdown(dataset)

    for column in dataset.columns:
        assert column.name in markdown
        assert column.description in markdown


# --------------------------------------------------------------------------- #
# Форматы читаются инструментами студента
# --------------------------------------------------------------------------- #


async def test_csv_is_plain_and_parses(session: AsyncSession) -> None:
    """Без комментариев в шапке: студенты открывают это Excel и Gretl,
    а те на строках с # спотыкаются."""
    summary = await seed(session)
    dataset = await build_round_dataset(session, summary.round_id)
    text = to_csv(dataset)

    assert not text.lstrip().startswith("#")

    rows = list(csv.DictReader(io.StringIO(text)))
    assert len(rows) == len(dataset.rows)
    assert set(rows[0]) == {c.name for c in dataset.columns}


async def test_xlsx_has_data_and_dictionary_sheets(session: AsyncSession) -> None:
    summary = await seed(session)
    dataset = await build_round_dataset(session, summary.round_id)

    book = load_workbook(io.BytesIO(to_xlsx(dataset)))
    assert book.sheetnames == ["Данные", "Словарь"]

    data_sheet = book["Данные"]
    assert data_sheet.max_row == len(dataset.rows) + 1  # + шапка


# --------------------------------------------------------------------------- #
# Воспроизводимость
# --------------------------------------------------------------------------- #


async def test_same_round_gives_same_dataset(session: AsyncSession) -> None:
    """Дважды скачали — получили одно и то же. Иначе разбор невозможен."""
    summary = await seed(session)
    first = to_csv(await build_round_dataset(session, summary.round_id))
    second = to_csv(await build_round_dataset(session, summary.round_id))
    assert first == second


async def test_missing_round_raises(session: AsyncSession) -> None:
    with pytest.raises(ValueError):
        await build_round_dataset(session, 999)


# --------------------------------------------------------------------------- #
# Данные под метод раунда
# --------------------------------------------------------------------------- #


async def _round_with_method(session: AsyncSession, method: Method) -> int:
    """Открытый раунд с заданным методом на том же рынке, что и пилот."""
    summary = await seed_oil_2013(session)
    source = await repo.get_round(session, summary.round_id)
    assert source is not None
    created = await repo.create_round(
        session,
        number=source.number + 1,
        method=method,
        difficulty=source.difficulty,
        market_a=source.market_a,
        market_b=source.market_b,
        market_mc=source.market_mc,
    )
    assert created.id is not None
    return created.id


async def test_regime_round_carries_its_own_column(session: AsyncSession) -> None:
    """Раунд по фиктивным переменным выдаёт столбец режима, а не голый спрос.

    До фазы 2 все шесть методов получали одну и ту же историю, и раунд по
    дамми ничем не отличался от раунда по парной регрессии — метод было
    нечего применять.
    """
    round_id = await _round_with_method(session, Method.OLS_MULTIPLE)
    dataset = await build_round_dataset(session, round_id)

    assert REGIME_COLUMN in {c.name for c in dataset.columns}
    flags = {row[REGIME_COLUMN] for row in dataset.rows}
    assert flags == {0.0, 1.0}

    # Столбец без описания бесполезен: студент не поймёт, что это за единица.
    assert REGIME_COLUMN in data_dictionary_markdown(dataset)

    # И доезжает до самого файла, а не только до объекта.
    rows = list(csv.DictReader(io.StringIO(to_csv(dataset))))
    assert {row[REGIME_COLUMN] for row in rows} == {"0.0", "1.0"}


async def test_round_without_a_case_refuses_to_export(session: AsyncSession) -> None:
    """Метод без кейса роняет выгрузку, а не подсовывает baseline.

    Громкий отказ на настройке раунда стоит минуты. Молчаливая подмена стоит
    раунда целиком: заявленному методу нечего было бы искать в данных,
    и заметили бы это последним.
    """
    missing = next((m for m in Method if m not in supported_methods()), None)
    if missing is None:
        pytest.skip("все шесть методов уже реализованы")

    round_id = await _round_with_method(session, missing)
    with pytest.raises(NotImplementedError):
        await build_round_dataset(session, round_id)
